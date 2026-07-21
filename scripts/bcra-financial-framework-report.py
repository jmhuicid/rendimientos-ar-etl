from __future__ import annotations

import argparse
import subprocess
import sys
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class VariableProfile:
    name: str
    block: str
    unit: str
    metric_kind: str
    preferred_frequency: str
    reading_hint: str


RATE_COLUMNS = {
    "Inflación mensual",
    "Inflación interanual",
    "REM",
    "TNA de depósitos a plazo fijo en pesos, 30-44 días",
    "TNA de depósitos a plazo fijo en pesos, 30-44 días, hasta $100.000",
    "TNA de depósitos a plazo fijo en pesos, 30-44 días, de más de $1.000.000",
    "TAMAR",
    "BADLAR",
}


PROFILES: dict[str, VariableProfile] = {
    "TC Minorista": VariableProfile("TC Minorista", "Tipo de cambio", "ARS/USD", "level", "diaria", "Referencia de precio minorista del dólar."),
    "TC Mayorista": VariableProfile("TC Mayorista", "Tipo de cambio", "ARS/USD", "level", "diaria", "Referencia mayorista y ancla para precios financieros."),
    "Base monetaria": VariableProfile("Base monetaria", "Agregados monetarios", "millones de pesos", "level", "semanal/mensual", "Pulso de expansión o absorción monetaria."),
    "Circulación monetaria": VariableProfile("Circulación monetaria", "Agregados monetarios", "millones de pesos", "level", "semanal/mensual", "Efectivo total en circulación."),
    "Billetes en público": VariableProfile("Billetes en público", "Agregados monetarios", "millones de pesos", "level", "semanal/mensual", "Demanda transaccional de efectivo del público."),
    "Efectivo en entidades financieras": VariableProfile("Efectivo en entidades financieras", "Agregados monetarios", "millones de pesos", "level", "semanal/mensual", "Liquidez física dentro del sistema financiero."),
    "Depósitos de bancos en cta. cte.": VariableProfile("Depósitos de bancos en cta. cte.", "Agregados monetarios", "millones de pesos", "level", "semanal/mensual", "Liquidez bancaria en cuenta corriente del BCRA."),
    "Depósitos en efectivo en entidades financieras": VariableProfile("Depósitos en efectivo en entidades financieras", "Depósitos", "millones de pesos", "level", "mensual", "Tamaño del fondeo bancario en efectivo."),
    "En cuentas corrientes": VariableProfile("En cuentas corrientes", "Depósitos", "millones de pesos", "level", "mensual", "Liquidez transaccional privada y pública."),
    "En Caja de ahorros": VariableProfile("En Caja de ahorros", "Depósitos", "millones de pesos", "level", "mensual", "Ahorro líquido de corto plazo."),
    "A plazo": VariableProfile("A plazo", "Depósitos", "millones de pesos", "level", "mensual", "Fondeo a plazo y preferencia por tasa."),
    "M2 privado": VariableProfile("M2 privado", "Agregados monetarios", "porcentaje", "rate", "diaria/semanal", "Lectura de liquidez privada amplia informada por BCRA."),
    "Préstamos de entidades financieras al sector privado": VariableProfile("Préstamos de entidades financieras al sector privado", "Crédito", "millones de pesos", "level", "mensual", "Pulso de crédito privado nominal."),
    "Inflación mensual": VariableProfile("Inflación mensual", "Inflación", "% mensual", "rate", "mensual", "Ritmo mensual observado de inflación."),
    "Inflación interanual": VariableProfile("Inflación interanual", "Inflación", "% interanual", "rate", "mensual", "Régimen inflacionario de doce meses."),
    "REM": VariableProfile("REM", "Expectativas", "%", "rate", "mensual", "Expectativa de inflación relevada por BCRA."),
    "TNA de depósitos a plazo fijo en pesos, 30-44 días": VariableProfile("TNA de depósitos a plazo fijo en pesos, 30-44 días", "Tasas", "% TNA", "rate", "diaria/semanal", "Tasa pasiva promedio de plazo fijo."),
    "TNA de depósitos a plazo fijo en pesos, 30-44 días, hasta $100.000": VariableProfile("TNA de depósitos a plazo fijo en pesos, 30-44 días, hasta $100.000", "Tasas", "% TNA", "rate", "diaria/semanal", "Tasa minorista aproximada de plazo fijo."),
    "TNA de depósitos a plazo fijo en pesos, 30-44 días, de más de $1.000.000": VariableProfile("TNA de depósitos a plazo fijo en pesos, 30-44 días, de más de $1.000.000", "Tasas", "% TNA", "rate", "diaria/semanal", "Tasa de mayor monto / tramo institucional."),
    "TAMAR": VariableProfile("TAMAR", "Tasas", "% TNA", "rate", "diaria/semanal", "Referencia mayorista de depósitos de mayor monto."),
    "BADLAR": VariableProfile("BADLAR", "Tasas", "% TNA", "rate", "diaria/semanal", "Referencia de depósitos mayoristas de bancos privados."),
    "CER": VariableProfile("CER", "Índices de inflación", "índice", "level", "diaria", "Coeficiente de estabilización de referencia para instrumentos CER."),
    "UVA": VariableProfile("UVA", "Índices de inflación", "índice", "level", "diaria", "Unidad ajustada por CER para crédito/ahorro."),
    "UVI": VariableProfile("UVI", "Índices de inflación", "índice", "level", "diaria", "Unidad indexada por costo de construcción/inflación según serie BCRA."),
}


def main() -> None:
    parser = argparse.ArgumentParser(description="Genera un marco financiero BCRA con evolución diaria/semanal/mensual/anual.")
    parser.add_argument("--today", default=date.today().isoformat(), help="Fecha de corrida YYYY-MM-DD.")
    parser.add_argument("--snapshot-dir", default=None, help="Directorio con bcra_public_series.xlsx/csv. Default: data/snapshots/YYYY-MM-DD.")
    parser.add_argument("--output-dir", default=None, help="Directorio de salida. Default: data/reports/YYYY-MM-DD.")
    parser.add_argument("--regenerate-snapshot", action="store_true", help="Fuerza regenerar bcra_public_series antes del reporte.")
    args = parser.parse_args()

    root = Path(__file__).resolve().parents[1]
    run_date = pd.to_datetime(args.today).date()
    snapshot_dir = Path(args.snapshot_dir) if args.snapshot_dir else root / "data" / "snapshots" / run_date.isoformat()
    output_dir = Path(args.output_dir) if args.output_dir else root / "data" / "reports" / run_date.isoformat()
    output_dir.mkdir(parents=True, exist_ok=True)

    snapshot_path = ensure_snapshot(root, snapshot_dir, run_date, args.regenerate_snapshot)
    series = load_series(snapshot_path)
    metadata = load_metadata(snapshot_path)

    daily = build_periodic_frame(series, "D")
    weekly = build_periodic_frame(series, "W-FRI")
    monthly = build_periodic_frame(series, "ME")
    annual = build_periodic_frame(series, "YE")
    summary = build_summary(series)
    changes = build_changes(series)
    block_moment = build_block_moment(summary)
    general_reading = build_general_reading(summary, block_moment, run_date, snapshot_path)
    framework = build_framework(summary, changes, run_date, snapshot_path)

    output_xlsx = output_dir / f"marco_financiero_bcra_{run_date:%Y%m%d}.xlsx"
    output_md = output_dir / f"marco_financiero_bcra_{run_date:%Y%m%d}.md"
    write_workbook(
        output_xlsx,
        {
            "Lectura_general": general_reading,
            "Marco_actual": framework,
            "Momento_por_bloque": block_moment,
            "Resumen_variables": summary,
            "Evolucion_diaria": daily,
            "Evolucion_semanal": weekly,
            "Evolucion_mensual": monthly,
            "Evolucion_anual": annual,
            "Cambios_por_variable": changes,
            "Metadata": metadata,
        },
    )
    write_markdown(output_md, general_reading, framework, block_moment, summary, changes, run_date, snapshot_path)

    print(f"OK Excel: {output_xlsx}")
    print(f"OK Markdown: {output_md}")


def ensure_snapshot(root: Path, snapshot_dir: Path, run_date: date, regenerate: bool) -> Path:
    xlsx_path = snapshot_dir / "bcra_public_series.xlsx"
    csv_path = snapshot_dir / "bcra_public_series.csv"
    if regenerate or not xlsx_path.exists():
        subprocess.run(
            [
                sys.executable,
                str(root / "scripts" / "bcra-public-series-export.py"),
                "--today",
                run_date.isoformat(),
                "--output-dir",
                str(snapshot_dir),
            ],
            check=True,
        )
    if xlsx_path.exists():
        return xlsx_path
    if csv_path.exists():
        return csv_path
    raise FileNotFoundError(f"No se encontró snapshot BCRA en {snapshot_dir}")


def load_series(path: Path) -> pd.DataFrame:
    if path.suffix.lower() == ".xlsx":
        frame = pd.read_excel(path, sheet_name="BCRA_public_series")
    else:
        frame = pd.read_csv(path)
    frame["Fecha"] = pd.to_datetime(frame["Fecha"], dayfirst=True, errors="coerce")
    frame = frame.dropna(subset=["Fecha"]).sort_values("Fecha").reset_index(drop=True)
    for column in frame.columns:
        if column != "Fecha":
            frame[column] = pd.to_numeric(frame[column], errors="coerce")
    return frame


def load_metadata(path: Path) -> pd.DataFrame:
    if path.suffix.lower() != ".xlsx":
        return pd.DataFrame()
    try:
        return pd.read_excel(path, sheet_name="Metadata")
    except ValueError:
        return pd.DataFrame()


def build_periodic_frame(series: pd.DataFrame, rule: str) -> pd.DataFrame:
    values = series.set_index("Fecha").sort_index()
    if rule == "D":
        out = values.tail(260).copy()
    else:
        out = values.resample(rule).last().dropna(how="all")
        out = out.tail({"W-FRI": 104, "ME": 60, "YE": 20}.get(rule, 100))
    return out.reset_index()


def build_summary(series: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    data = series.set_index("Fecha").sort_index()
    for column in [c for c in series.columns if c != "Fecha"]:
        values = data[column].dropna()
        if values.empty:
            continue
        profile = PROFILES.get(column, default_profile(column))
        latest_date = values.index[-1]
        latest_value = float(values.iloc[-1])
        row = {
            "bloque": profile.block,
            "variable": column,
            "unidad": profile.unit,
            "tipo_metrica": "tasa" if profile.metric_kind == "rate" else "nivel/indice",
            "frecuencia_sugerida": profile.preferred_frequency,
            "ultima_fecha": latest_date.date().isoformat(),
            "ultimo_valor": latest_value,
            "cambio_dato_previo": diff_from_previous(values),
            "var_7d": period_change(values, latest_date, 7, profile.metric_kind),
            "var_30d": period_change(values, latest_date, 30, profile.metric_kind),
            "var_90d": period_change(values, latest_date, 90, profile.metric_kind),
            "var_ytd": ytd_change(values, latest_date, profile.metric_kind),
            "var_1y": period_change(values, latest_date, 365, profile.metric_kind),
            "lectura": profile.reading_hint,
        }
        rows.append(row)
    return pd.DataFrame(rows).sort_values(["bloque", "variable"]).reset_index(drop=True)


def build_changes(series: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    data = series.set_index("Fecha").sort_index()
    windows = [("7d", 7), ("30d", 30), ("90d", 90), ("1y", 365)]
    for column in [c for c in series.columns if c != "Fecha"]:
        values = data[column].dropna()
        if values.empty:
            continue
        profile = PROFILES.get(column, default_profile(column))
        latest_date = values.index[-1]
        latest_value = float(values.iloc[-1])
        for label, days in windows:
            previous_date = latest_date - pd.Timedelta(days=days)
            previous_value = value_on_or_before(values, previous_date)
            if previous_value is None:
                continue
            rows.append(
                {
                    "bloque": profile.block,
                    "variable": column,
                    "ventana": label,
                    "fecha_actual": latest_date.date().isoformat(),
                    "valor_actual": latest_value,
                    "fecha_base": values[values.index <= previous_date].index[-1].date().isoformat(),
                    "valor_base": float(previous_value),
                    "cambio_abs": latest_value - float(previous_value),
                    "cambio_pct_o_pp": period_change(values, latest_date, days, profile.metric_kind),
                    "unidad_cambio": "p.p." if profile.metric_kind == "rate" else "%",
                }
            )
    return pd.DataFrame(rows)


def build_framework(summary: pd.DataFrame, changes: pd.DataFrame, run_date: date, snapshot_path: Path) -> pd.DataFrame:
    rows = [
        {
            "seccion": "Corte",
            "metrica": "Fecha de generación",
            "valor": run_date.isoformat(),
            "lectura": f"Reporte construido con {snapshot_path.name}; marco de trabajo para Finanzas con series públicas BCRA disponibles.",
        },
        {
            "seccion": "Cobertura",
            "metrica": "Variables disponibles",
            "valor": len(summary),
            "lectura": "Incluye tipo de cambio, agregados monetarios, depósitos, crédito privado, inflación, expectativas, tasas e índices CER/UVA/UVI.",
        },
    ]
    rows.extend(block_readings(summary))
    rows.extend(block_moment_readings(summary))
    rows.extend(relative_rate_readings(summary))
    rows.extend(momentum_readings(summary))
    return pd.DataFrame(rows)


def build_general_reading(summary: pd.DataFrame, block_moment: pd.DataFrame, run_date: date, snapshot_path: Path) -> pd.DataFrame:
    states = {str(row["bloque"]): str(row["estado"]) for _, row in block_moment.iterrows()}
    rows = [
        {
            "orden": 1,
            "seccion": "Diagnostico general",
            "titulo": "Regimen nominal actual",
            "lectura": general_market_diagnosis(summary, states),
            "implicancia_finanzas": "Usar el tablero como marco diario para decidir posicionamiento en pesos, cobertura, duration bancaria e instrumentos indexados.",
            "senal_a_monitorear": "Cambio simultaneo de tres ejes: aceleracion cambiaria, tasas reales ex-ante y velocidad de indexacion CER/UVA/UVI.",
        },
        {
            "orden": 2,
            "seccion": "Balance de fuerzas",
            "titulo": "Tasas versus indexacion",
            "lectura": rates_vs_indexation_reading(summary),
            "implicancia_finanzas": "Comparar retorno nominal reinvertible contra ajuste por inflacion para no confundir tasa alta con rendimiento real positivo.",
            "senal_a_monitorear": "Si CER/UVA/UVI a 30 dias anualizado o mensualizado supera persistentemente a BADLAR/TAMAR efectiva, aumenta el valor relativo de cobertura indexada.",
        },
        {
            "orden": 3,
            "seccion": "Liquidez y fondeo",
            "titulo": "Preferencia por liquidez",
            "lectura": liquidity_funding_reading(summary, states),
            "implicancia_finanzas": "Una suba de liquidez transaccional con plazo fijo lateral puede anticipar menor apetito por inmovilizar pesos.",
            "senal_a_monitorear": "Caja de ahorro y cuentas corrientes creciendo por encima de depositos a plazo durante varias semanas.",
        },
        {
            "orden": 4,
            "seccion": "Riesgo de mercado",
            "titulo": "Tipo de cambio como ancla",
            "lectura": fx_risk_reading(summary, states),
            "implicancia_finanzas": "El tipo de cambio mayorista ordena precios relativos de activos locales, bonos dollar-linked y expectativas de cobertura.",
            "senal_a_monitorear": "Salto de TC mayorista 30d por encima de la velocidad de tasas pasivas o de inflacion mensual observada.",
        },
        {
            "orden": 5,
            "seccion": "Trabajo diario",
            "titulo": "Uso operativo recomendado",
            "lectura": "Leer primero la hoja Marco_actual, luego Momento_por_bloque y finalmente Cambios_por_variable para identificar que variable explica el movimiento.",
            "implicancia_finanzas": "Permite separar ruido diario de cambio de regimen y priorizar conversaciones de tasas, liquidez, cobertura e indexacion.",
            "senal_a_monitorear": "Nuevos maximos de movimiento 30d, revisiones BCRA y desalineaciones entre expectativas, inflacion e instrumentos indexados.",
        },
    ]
    rows.extend(
        {
            "orden": 10 + idx,
            "seccion": "Momento por bloque",
            "titulo": row["bloque"],
            "lectura": row["lectura_momento"],
            "implicancia_finanzas": row["implicancia_finanzas"],
            "senal_a_monitorear": row["senal_a_monitorear"],
        }
        for idx, row in block_moment.reset_index(drop=True).iterrows()
    )
    rows.append(
        {
            "orden": 99,
            "seccion": "Trazabilidad",
            "titulo": "Fuente y corte",
            "lectura": f"Fecha de generacion {run_date.isoformat()}, construido desde {snapshot_path.name}.",
            "implicancia_finanzas": "La lectura respeta la ultima fecha disponible por serie; no fuerza imputaciones para series con rezagos de publicacion.",
            "senal_a_monitorear": "Diferencias de fecha entre bloques: tasas, monetarias e inflacion pueden tener rezagos distintos.",
        }
    )
    return pd.DataFrame(rows)


def general_market_diagnosis(summary: pd.DataFrame, states: dict[str, str]) -> str:
    fx = states.get("Tipo de cambio", "sin lectura")
    rates = states.get("Tasas", "sin lectura")
    inflation = states.get("Inflacion", states.get("Inflación", "sin lectura"))
    liquidity = states.get("Agregados monetarios", "sin lectura")
    deposits = states.get("Depositos", states.get("Depósitos", "sin lectura"))
    credit = states.get("Credito", states.get("Crédito", "sin lectura"))
    indexation = states.get("Indices de inflacion", states.get("Índices de inflación", "sin lectura"))
    return (
        f"El marco muestra {fx} en tipo de cambio, {rates} en tasas, {inflation} en inflacion, "
        f"{indexation} en indexacion, {liquidity} en liquidez, {deposits} en depositos y {credit} en credito. "
        "La foto sugiere un mercado todavia nominal, pero con desinflacion operativa y tasas que no aceleran al mismo ritmo que la liquidez transaccional."
    )


def rates_vs_indexation_reading(summary: pd.DataFrame) -> str:
    badlar = first_available(summary, "BADLAR", "ultimo_valor")
    tamar = first_available(summary, "TAMAR", "ultimo_valor")
    uva_30d = first_available(summary, "UVA", "var_30d")
    uvi_30d = first_available(summary, "UVI", "var_30d")
    inflation = first_available(summary, "Inflación mensual", "ultimo_valor")
    return (
        f"BADLAR/TAMAR estan en {format_inline(badlar)}%/{format_inline(tamar)}% TNA. "
        f"UVA avanza {format_inline(uva_30d)}% y UVI {format_inline(uvi_30d)}% en 30 dias, con inflacion mensual en {format_inline(inflation)}%. "
        "La comparacion relevante para Finanzas es si la tasa pasiva compensa la indexacion efectiva y el costo de liquidez."
    )


def liquidity_funding_reading(summary: pd.DataFrame, states: dict[str, str]) -> str:
    base = first_available(summary, "Base monetaria", "var_30d")
    caja = first_available(summary, "En Caja de ahorros", "var_30d")
    plazo = first_available(summary, "A plazo", "var_30d")
    return (
        f"El bloque monetario aparece como {states.get('Agregados monetarios', 'sin lectura')}; "
        f"base monetaria sube {format_inline(base)}% en 30 dias. En depositos, caja de ahorro cambia {format_inline(caja)}% "
        f"contra plazo fijo {format_inline(plazo)}%, senal de preferencia relativa por liquidez cuando la brecha favorece saldos transaccionales."
    )


def fx_risk_reading(summary: pd.DataFrame, states: dict[str, str]) -> str:
    mayorista = first_available(summary, "TC Mayorista", "ultimo_valor")
    fx_30d = first_available(summary, "TC Mayorista", "var_30d")
    fx_1y = first_available(summary, "TC Mayorista", "var_1y")
    return (
        f"El tipo de cambio esta en estado {states.get('Tipo de cambio', 'sin lectura')}: mayorista {format_inline(mayorista)}, "
        f"{format_inline(fx_30d)}% en 30 dias y {format_inline(fx_1y)}% interanual. "
        "Mientras el movimiento sea moderado, el foco pasa a spread de tasas e instrumentos indexados; si acelera, cambia la prioridad hacia cobertura."
    )


def block_readings(summary: pd.DataFrame) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for block, group in summary.groupby("bloque"):
        latest_dates = sorted(group["ultima_fecha"].dropna().astype(str).unique())
        rows.append(
            {
                "seccion": "Bloques disponibles",
                "metrica": block,
                "valor": len(group),
                "lectura": f"{len(group)} variables; última fecha observada {latest_dates[-1] if latest_dates else 'n/d'}.",
            }
        )
    return rows


def build_block_moment(summary: pd.DataFrame) -> pd.DataFrame:
    return pd.DataFrame(block_moment_rows(summary))


def block_moment_readings(summary: pd.DataFrame) -> list[dict[str, Any]]:
    return [
        {
            "seccion": "Momento de mercado",
            "metrica": row["bloque"],
            "valor": row["estado"],
            "lectura": row["lectura_momento"],
        }
        for row in block_moment_rows(summary)
    ]


def block_moment_rows(summary: pd.DataFrame) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for block in sorted(summary["bloque"].dropna().unique()):
        group = summary[summary["bloque"] == block]
        rows.append(
            {
                "bloque": block,
                "estado": block_state(block, summary),
                "lectura_momento": block_market_reading(block, summary),
                "implicancia_finanzas": block_finance_implication(block, summary),
                "senal_a_monitorear": block_monitoring_signal(block, summary),
                "variable_lider_30d": leading_variable(group),
                "movimiento_lider_30d": leading_value(group),
                "ultima_fecha_bloque": latest_block_date(group),
                "variables": len(group),
            }
        )
    return rows


def block_finance_implication(block: str, summary: pd.DataFrame) -> str:
    if block == "Tipo de cambio":
        return "Define sensibilidad de cobertura, valuacion de instrumentos hard-dollar/dollar-linked y traslado esperado a precios nominales."
    if block == "Tasas":
        return "Ordena costo de oportunidad de caja en pesos, renovacion de plazos fijos y comparacion contra instrumentos CER/UVA."
    if block == "Inflación":
        return "Determina rendimiento real minimo exigido y la urgencia de cobertura indexada."
    if block == "Expectativas":
        return "Sirve como ancla prospectiva para presupuesto financiero, tasa requerida y escenarios de desinflacion."
    if block == "Índices de inflación":
        return "Mide la capitalizacion efectiva de indexados y permite contrastar tasa nominal contra ajuste observado."
    if block == "Agregados monetarios":
        return "Ayuda a anticipar presion nominal, disponibilidad de pesos y liquidez del sistema."
    if block == "Depósitos":
        return "Muestra si el mercado prefiere liquidez inmediata o inmovilizar pesos a tasa."
    if block == "Crédito":
        return "Aporta lectura de actividad nominal, demanda de financiamiento y expansion del balance privado."
    return "Bloque disponible para seguimiento y contraste con el resto del tablero."


def block_monitoring_signal(block: str, summary: pd.DataFrame) -> str:
    if block == "Tipo de cambio":
        return "Aceleracion del TC mayorista 30d por encima de tasas pasivas o de la indexacion mensual."
    if block == "Tasas":
        return "Cambio de pendiente BADLAR/TAMAR y spread TAMAR-BADLAR; caidas con inflacion/indexacion firme deterioran tasa real."
    if block == "Inflación":
        return "Reversion de inflacion mensual o freno en la baja interanual."
    if block == "Expectativas":
        return "Suba del REM o divergencia entre REM, inflacion observada y tasa de mercado."
    if block == "Índices de inflación":
        return "Aceleracion simultanea de CER/UVA/UVI en 30d y 90d."
    if block == "Agregados monetarios":
        return "Base/circulacion creciendo mas rapido que tasas e inflacion esperada."
    if block == "Depósitos":
        return "Caja de ahorro/cuentas corrientes creciendo sobre plazo fijo, o salida de depositos a plazo."
    if block == "Crédito":
        return "Credito nominal desacelerando con liquidez alta, o acelerando por encima de fondeo estable."
    return "Cambios de 30d y 90d fuera del rango reciente."


def block_state(block: str, summary: pd.DataFrame) -> str:
    if block == "Tipo de cambio":
        move = first_available(summary, "TC Mayorista", "var_30d")
        if move is None:
            return "sin lectura"
        if move >= 5:
            return "presion cambiaria"
        if move >= 1:
            return "deslizamiento moderado"
        if move <= -1:
            return "apreciacion/compresion"
        return "estabilidad relativa"
    if block == "Tasas":
        badlar = first_available(summary, "BADLAR", "var_30d")
        tamar = first_available(summary, "TAMAR", "var_30d")
        avg_move = mean_available([badlar, tamar])
        if avg_move is None:
            return "sin lectura"
        if avg_move >= 1:
            return "tasas en suba"
        if avg_move <= -1:
            return "tasas en baja"
        return "tasas laterales"
    if block == "Inflación":
        monthly = first_available(summary, "Inflación mensual", "ultimo_valor")
        yearly_move = first_available(summary, "Inflación interanual", "var_1y")
        if monthly is None:
            return "sin lectura"
        if monthly <= 2 and (yearly_move is None or yearly_move <= 0):
            return "desinflacion operativa"
        if monthly >= 3:
            return "inflacion elevada"
        return "inflacion en monitoreo"
    if block == "Expectativas":
        rem = first_available(summary, "REM", "var_30d")
        if rem is None:
            return "sin lectura"
        if rem < 0:
            return "expectativas mejorando"
        if rem > 0:
            return "expectativas deteriorando"
        return "expectativas estables"
    if block == "Índices de inflación":
        uva_30d = first_available(summary, "UVA", "var_30d")
        if uva_30d is None:
            return "sin lectura"
        if uva_30d >= 3:
            return "indexacion acelerada"
        if uva_30d <= 2:
            return "indexacion moderada"
        return "indexacion intermedia"
    if block == "Agregados monetarios":
        base = first_available(summary, "Base monetaria", "var_30d")
        circulation = first_available(summary, "Circulación monetaria", "var_30d")
        avg_move = mean_available([base, circulation])
        if avg_move is None:
            return "sin lectura"
        if avg_move >= 5:
            return "liquidez expandiendose"
        if avg_move <= 0:
            return "liquidez contenida"
        return "liquidez creciendo moderada"
    if block == "Depósitos":
        plazo = first_available(summary, "A plazo", "var_30d")
        caja = first_available(summary, "En Caja de ahorros", "var_30d")
        if plazo is None and caja is None:
            return "sin lectura"
        if plazo is not None and caja is not None and caja > plazo + 3:
            return "preferencia por liquidez"
        if plazo is not None and plazo > 3:
            return "fondeo a plazo creciendo"
        return "depositos mixtos/laterales"
    if block == "Crédito":
        credit = first_available(summary, "Préstamos de entidades financieras al sector privado", "var_30d")
        if credit is None:
            return "sin lectura"
        if credit >= 3:
            return "credito nominal creciendo"
        if credit <= 0:
            return "credito nominal frenado"
        return "credito en expansion moderada"
    return "seguimiento disponible"


def block_market_reading(block: str, summary: pd.DataFrame) -> str:
    if block == "Tipo de cambio":
        mayorista = first_available(summary, "TC Mayorista", "ultimo_valor")
        move_30d = first_available(summary, "TC Mayorista", "var_30d")
        move_1y = first_available(summary, "TC Mayorista", "var_1y")
        return f"El mayorista opera en {format_inline(mayorista)}; en 30 dias se movio {format_inline(move_30d)}%, y en un año {format_inline(move_1y)}%. Sirve como ancla para precios de activos locales y cobertura."
    if block == "Tasas":
        badlar = first_available(summary, "BADLAR", "ultimo_valor")
        tamar = first_available(summary, "TAMAR", "ultimo_valor")
        spread = None if badlar is None or tamar is None else tamar - badlar
        return f"BADLAR esta en {format_inline(badlar)}% TNA y TAMAR en {format_inline(tamar)}% TNA; el spread TAMAR-BADLAR es {format_inline(spread)} p.p. El bloque marca el costo de oportunidad de pesos."
    if block == "Inflación":
        monthly = first_available(summary, "Inflación mensual", "ultimo_valor")
        annual = first_available(summary, "Inflación interanual", "ultimo_valor")
        monthly_move = first_available(summary, "Inflación mensual", "var_30d")
        return f"Inflacion mensual en {format_inline(monthly)}% e interanual en {format_inline(annual)}%; el cambio mensual fue {format_inline(monthly_move)} p.p. Define el piso para rendimiento real requerido."
    if block == "Expectativas":
        rem = first_available(summary, "REM", "ultimo_valor")
        rem_30d = first_available(summary, "REM", "var_30d")
        return f"REM en {format_inline(rem)}%, con variacion de 30 dias de {format_inline(rem_30d)} p.p. Resume el sesgo esperado por el mercado para precios nominales."
    if block == "Índices de inflación":
        cer = first_available(summary, "CER", "var_30d")
        uva = first_available(summary, "UVA", "var_30d")
        uvi = first_available(summary, "UVI", "var_30d")
        return f"CER/UVA/UVI avanzan {format_inline(cer)}%, {format_inline(uva)}% y {format_inline(uvi)}% en 30 dias. Es la referencia para instrumentos indexados y comparacion contra tasas nominales."
    if block == "Agregados monetarios":
        base = first_available(summary, "Base monetaria", "var_30d")
        circulation = first_available(summary, "Circulación monetaria", "var_30d")
        m2 = first_available(summary, "M2 privado", "var_30d")
        return f"Base monetaria cambia {format_inline(base)}%, circulacion {format_inline(circulation)}% y M2 privado {format_inline(m2)} p.p. en 30 dias. Mide liquidez disponible y presion nominal potencial."
    if block == "Depósitos":
        plazo = first_available(summary, "A plazo", "var_30d")
        caja = first_available(summary, "En Caja de ahorros", "var_30d")
        ctas = first_available(summary, "En cuentas corrientes", "var_30d")
        return f"A plazo cambia {format_inline(plazo)}%, caja de ahorro {format_inline(caja)}% y cuentas corrientes {format_inline(ctas)}% en 30 dias. Indica preferencia por liquidez versus duration bancaria."
    if block == "Crédito":
        credit = first_available(summary, "Préstamos de entidades financieras al sector privado", "var_30d")
        credit_yoy = first_available(summary, "Préstamos de entidades financieras al sector privado", "var_1y")
        return f"Prestamos al sector privado crecen {format_inline(credit)}% en 30 dias y {format_inline(credit_yoy)}% interanual nominal. Ayuda a leer actividad, demanda de pesos y apalancamiento."
    latest = latest_block_date(summary[summary["bloque"] == block])
    return f"Bloque disponible para seguimiento; ultima fecha observada {latest}."


def first_available(summary: pd.DataFrame, variable: str, column: str) -> float | None:
    row = summary[summary["variable"] == variable]
    if row.empty or column not in row:
        return None
    value = row.iloc[0][column]
    if pd.isna(value):
        return None
    return float(value)


def mean_available(values: list[float | None]) -> float | None:
    valid = [value for value in values if value is not None]
    if not valid:
        return None
    return sum(valid) / len(valid)


def leading_variable(group: pd.DataFrame) -> str:
    valid = group.dropna(subset=["var_30d"]).copy()
    if valid.empty:
        return ""
    idx = valid["var_30d"].abs().idxmax()
    return str(valid.loc[idx, "variable"])


def leading_value(group: pd.DataFrame) -> float | None:
    valid = group.dropna(subset=["var_30d"]).copy()
    if valid.empty:
        return None
    idx = valid["var_30d"].abs().idxmax()
    return float(valid.loc[idx, "var_30d"])


def latest_block_date(group: pd.DataFrame) -> str:
    dates = sorted(group["ultima_fecha"].dropna().astype(str).unique())
    return dates[-1] if dates else ""


def format_inline(value: float | None) -> str:
    if value is None or pd.isna(value):
        return "n/d"
    return f"{float(value):.2f}"


def relative_rate_readings(summary: pd.DataFrame) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    latest = summary.set_index("variable")
    for left, right in [
        ("TAMAR", "BADLAR"),
        ("TNA de depósitos a plazo fijo en pesos, 30-44 días", "BADLAR"),
        ("Inflación mensual", "REM"),
    ]:
        if left in latest.index and right in latest.index:
            spread = float(latest.loc[left, "ultimo_valor"]) - float(latest.loc[right, "ultimo_valor"])
            rows.append(
                {
                    "seccion": "Relaciones clave",
                    "metrica": f"{left} vs {right}",
                    "valor": round(spread, 4),
                    "lectura": f"Diferencial actual de {spread:.2f} puntos; útil para monitorear premio relativo y expectativas.",
                }
            )
    if "UVI" in latest.index and "UVA" in latest.index:
        ratio = float(latest.loc["UVI", "ultimo_valor"]) / float(latest.loc["UVA", "ultimo_valor"])
        rows.append(
            {
                "seccion": "Relaciones clave",
                "metrica": "UVI / UVA",
                "valor": round(ratio, 4),
                "lectura": "Relación entre unidades indexadas para seguir divergencias de indexación.",
            }
        )
    return rows


def momentum_readings(summary: pd.DataFrame) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    candidates = summary.dropna(subset=["var_30d"]).copy()
    if candidates.empty:
        return rows
    strongest = candidates.reindex(candidates["var_30d"].abs().sort_values(ascending=False).index).head(5)
    for _, row in strongest.iterrows():
        unit = "p.p." if row["tipo_metrica"] == "tasa" else "%"
        rows.append(
            {
                "seccion": "Movimientos 30d destacados",
                "metrica": row["variable"],
                "valor": round(float(row["var_30d"]), 4),
                "lectura": f"Movimiento de 30 días: {float(row['var_30d']):.2f} {unit}.",
            }
        )
    return rows


def diff_from_previous(values: pd.Series) -> float | None:
    if len(values) < 2:
        return None
    return float(values.iloc[-1] - values.iloc[-2])


def period_change(values: pd.Series, latest_date: pd.Timestamp, days: int, metric_kind: str) -> float | None:
    previous = value_on_or_before(values, latest_date - pd.Timedelta(days=days))
    if previous is None:
        return None
    current = float(values.iloc[-1])
    previous = float(previous)
    if metric_kind == "rate":
        return current - previous
    if previous == 0:
        return None
    return (current / previous - 1) * 100


def ytd_change(values: pd.Series, latest_date: pd.Timestamp, metric_kind: str) -> float | None:
    start = pd.Timestamp(year=latest_date.year, month=1, day=1)
    previous = value_on_or_before(values, start)
    if previous is None:
        return None
    current = float(values.iloc[-1])
    previous = float(previous)
    if metric_kind == "rate":
        return current - previous
    if previous == 0:
        return None
    return (current / previous - 1) * 100


def value_on_or_before(values: pd.Series, target: pd.Timestamp) -> float | None:
    prior = values[values.index <= target].dropna()
    if prior.empty:
        return None
    return float(prior.iloc[-1])


def default_profile(column: str) -> VariableProfile:
    kind = "rate" if column in RATE_COLUMNS else "level"
    return VariableProfile(column, "Otros", "", kind, "según disponibilidad", "Serie disponible para seguimiento financiero.")


def write_workbook(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=name[:31], index=False)

    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        for column_cells in ws.columns:
            width = min(max(len(str(c.value)) if c.value is not None else 0 for c in column_cells) + 2, 60)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width
        ws.auto_filter.ref = ws.dimensions

    add_line_chart(wb, "Evolucion_mensual", "BCRA mensual - variables seleccionadas", ["TC Mayorista", "Base monetaria", "BADLAR", "TAMAR", "CER", "UVA", "UVI"], "J2")
    add_line_chart(wb, "Evolucion_anual", "BCRA anual - cierre de año", ["TC Mayorista", "Base monetaria", "BADLAR", "TAMAR", "CER", "UVA", "UVI"], "J2")
    wb.save(path)


def add_line_chart(wb, sheet_name: str, title: str, desired_columns: list[str], anchor: str) -> None:
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    if ws.max_row < 3:
        return
    header = [cell.value for cell in ws[1]]
    selected = [header.index(col) + 1 for col in desired_columns if col in header]
    if not selected:
        return
    chart = LineChart()
    chart.title = title
    chart.y_axis.title = "Nivel / tasa"
    chart.x_axis.title = "Fecha"
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    for col_idx in selected[:7]:
        data = Reference(ws, min_col=col_idx, max_col=col_idx, min_row=1, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 9
    chart.width = 20
    ws.add_chart(chart, anchor)


def write_markdown(
    path: Path,
    general_reading: pd.DataFrame,
    framework: pd.DataFrame,
    block_moment: pd.DataFrame,
    summary: pd.DataFrame,
    changes: pd.DataFrame,
    run_date: date,
    snapshot_path: Path,
) -> None:
    lines = [
        "# Marco financiero BCRA",
        "",
        f"Fecha de generación: {run_date.isoformat()}",
        f"Snapshot fuente: `{snapshot_path}`",
        "",
        "## Lectura general",
        "",
    ]
    for _, row in general_reading.iterrows():
        lines.append(f"### {row['titulo']}")
        lines.append(str(row["lectura"]))
        lines.append("")
        lines.append(f"**Implicancia para Finanzas:** {row['implicancia_finanzas']}")
        lines.append("")
        lines.append(f"**Señal a monitorear:** {row['senal_a_monitorear']}")
        lines.append("")

    lines.extend(["## Marco actual", ""])
    for _, row in framework.iterrows():
        lines.append(f"- **{row['seccion']} / {row['metrica']}**: {row['valor']}. {row['lectura']}")

    lines.extend(["", "## Momento por bloque", ""])
    block_display = block_moment.copy()
    if "movimiento_lider_30d" in block_display:
        block_display["movimiento_lider_30d"] = block_display["movimiento_lider_30d"].map(format_number)
    lines.extend(markdown_table(block_display))

    lines.extend(["", "## Último dato por variable", ""])
    display = summary.copy()
    numeric_cols = ["ultimo_valor", "cambio_dato_previo", "var_7d", "var_30d", "var_90d", "var_ytd", "var_1y"]
    for col in numeric_cols:
        if col in display:
            display[col] = display[col].map(format_number)
    lines.extend(markdown_table(display))

    lines.extend(["", "## Cambios por ventana", ""])
    if not changes.empty:
        changes_display = changes.copy()
        for col in ["valor_actual", "valor_base", "cambio_abs", "cambio_pct_o_pp"]:
            changes_display[col] = changes_display[col].map(format_number)
        lines.extend(markdown_table(changes_display.head(120)))
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def markdown_table(frame: pd.DataFrame) -> list[str]:
    columns = list(frame.columns)
    lines = [
        "| " + " | ".join(columns) + " |",
        "| " + " | ".join("---" for _ in columns) + " |",
    ]
    for _, row in frame.iterrows():
        values = ["" if pd.isna(row[col]) else str(row[col]).replace("|", "\\|") for col in columns]
        lines.append("| " + " | ".join(values) + " |")
    return lines


def format_number(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    return f"{float(value):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


if __name__ == "__main__":
    main()
