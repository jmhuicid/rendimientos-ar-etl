from __future__ import annotations

import argparse
import json
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any
from urllib.parse import urlencode
from urllib.request import Request, urlopen

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


BCRA_BASE_URL = "https://api.bcra.gob.ar/estadisticas/v4.0/Monetarias"


@dataclass(frozen=True)
class BcraColumn:
    variable_id: int
    output_name: str
    source_name: str | None = None
    note: str | None = None


BCRA_PUBLIC_SERIES_COLUMNS: tuple[BcraColumn, ...] = (
    BcraColumn(4, "TC Minorista"),
    BcraColumn(5, "TC Mayorista"),
    BcraColumn(15, "Base monetaria"),
    BcraColumn(16, "Circulación monetaria"),
    BcraColumn(17, "Billetes en público"),
    BcraColumn(18, "Efectivo en entidades financieras"),
    BcraColumn(19, "Depósitos de bancos en cta. cte."),
    BcraColumn(21, "Depósitos en efectivo en entidades financieras"),
    BcraColumn(22, "En cuentas corrientes"),
    BcraColumn(23, "En Caja de ahorros"),
    BcraColumn(24, "A plazo"),
    BcraColumn(25, "M2 privado"),
    BcraColumn(26, "Préstamos de entidades financieras al sector privado"),
    BcraColumn(27, "Inflación mensual"),
    BcraColumn(28, "Inflación interanual"),
    BcraColumn(29, "REM"),
    BcraColumn(
        1189,
        "TNA de depósitos a plazo fijo en pesos, 30-44 días",
        source_name="Tasa de interés de depósitos a plazo fijo en pesos",
        note="Reemplaza al ID histórico 128, no disponible en BCRA v4.",
    ),
    BcraColumn(
        1190,
        "TNA de depósitos a plazo fijo en pesos, 30-44 días, hasta $100.000",
        source_name="Tasa de interés de depósitos a plazo fijo en pesos de personas humanas",
        note="Reemplaza al ID histórico 129, no disponible en BCRA v4.",
    ),
    BcraColumn(
        1192,
        "TNA de depósitos a plazo fijo en pesos, 30-44 días, de más de $1.000.000",
        source_name="Tasa de interés de depósitos a plazo fijo en pesos de otras personas jurídicas",
        note="Reemplaza al ID histórico 131, no disponible en BCRA v4.",
    ),
    BcraColumn(136, "TAMAR"),
    BcraColumn(139, "BADLAR"),
    BcraColumn(30, "CER"),
    BcraColumn(31, "UVA"),
    BcraColumn(32, "UVI"),
)


def main() -> None:
    parser = argparse.ArgumentParser(description="Exporta series públicas BCRA para rendimientos-ar-etl.")
    parser.add_argument("--today", default=date.today().isoformat(), help="Fecha de corrida YYYY-MM-DD.")
    parser.add_argument("--output-root", default=None, help="Directorio base. Default: data/snapshots.")
    parser.add_argument("--output-dir", default=None, help="Directorio final. Si se omite usa data/snapshots/YYYY-MM-DD.")
    parser.add_argument("--limit", type=int, default=1000, help="Tamaño de página para BCRA v4.")
    args = parser.parse_args()

    root = Path(__file__).resolve().parents[1]
    run_date = pd.to_datetime(args.today).date()
    output_dir = (
        Path(args.output_dir)
        if args.output_dir
        else (Path(args.output_root) if args.output_root else root / "data" / "snapshots") / run_date.isoformat()
    )
    output_dir.mkdir(parents=True, exist_ok=True)

    frame, metadata = build_public_series_frame(as_of=run_date, limit=args.limit)
    csv_path = output_dir / "bcra_public_series.csv"
    xlsx_path = output_dir / "bcra_public_series.xlsx"
    frame.to_csv(csv_path, index=False, encoding="utf-8-sig")
    write_xlsx(xlsx_path, frame, metadata)
    update_snapshot_manifest(root, output_dir, csv_path, xlsx_path, len(frame))

    print(f"OK CSV: {csv_path}")
    print(f"OK Excel: {xlsx_path}")
    print(f"Filas: {len(frame)}")


def build_public_series_frame(as_of: date, limit: int = 1000) -> tuple[pd.DataFrame, pd.DataFrame]:
    frames: list[pd.DataFrame] = []
    metadata_rows: list[dict[str, Any]] = []
    catalog = fetch_catalog_metadata()

    for column in BCRA_PUBLIC_SERIES_COLUMNS:
        payload = fetch_bcra_series(column.variable_id, catalog, limit=limit)
        detail = payload["detail"]
        meta = payload["metadata"]
        metadata_rows.append(
            {
                "idVariable": column.variable_id,
                "columna_salida": column.output_name,
                "descripcion_fuente": column.source_name or meta.get("descripcion") or column.output_name,
                "primer_fecha_catalogo": meta.get("primerFechaInformada"),
                "ultima_fecha_catalogo": meta.get("ultFechaInformada"),
                "ultimo_valor_catalogo": meta.get("ultValorInformado"),
                "nota": column.note or "",
            }
        )
        if not detail:
            continue
        frame = pd.DataFrame(detail)
        if frame.empty or "fecha" not in frame or "valor" not in frame:
            continue
        frame = frame[["fecha", "valor"]].copy()
        frame["Fecha"] = pd.to_datetime(frame["fecha"], errors="coerce")
        frame[column.output_name] = pd.to_numeric(frame["valor"], errors="coerce")
        frame = frame.dropna(subset=["Fecha"])[["Fecha", column.output_name]]
        frame = frame[frame["Fecha"].dt.date <= as_of]
        frames.append(frame)

    output_columns = ["Fecha", *[item.output_name for item in BCRA_PUBLIC_SERIES_COLUMNS]]
    if frames:
        out = frames[0]
        for frame in frames[1:]:
            out = out.merge(frame, on="Fecha", how="outer")
        out = out.sort_values("Fecha", ascending=False).reset_index(drop=True)
        out = out.reindex(columns=output_columns)
        out["Fecha"] = out["Fecha"].dt.strftime("%d/%m/%Y")
    else:
        out = pd.DataFrame(columns=output_columns)

    return out, pd.DataFrame(metadata_rows)


def fetch_bcra_series(variable_id: int, catalog: dict[int, dict[str, Any]], limit: int = 1000) -> dict[str, Any]:
    rows: list[dict[str, Any]] = []
    offset = 0
    total: int | None = None
    latest_payload: dict[str, Any] | None = None
    while total is None or offset < total:
        query = urlencode({"limit": max(int(limit), 1), "offset": offset})
        payload = fetch_json(f"{BCRA_BASE_URL}/{int(variable_id)}?{query}")
        latest_payload = payload
        resultset = payload.get("metadata", {}).get("resultset", {})
        total = int(resultset.get("count", 0))
        detail = payload.get("results", [{}])[0].get("detalle", [])
        if not detail:
            break
        rows.extend(detail)
        offset += max(int(limit), 1)
    catalog_meta = catalog.get(int(variable_id), {})
    return {"detail": rows, "metadata": catalog_meta, "raw": latest_payload or {}}


def fetch_catalog_metadata() -> dict[int, dict[str, Any]]:
    payload = fetch_json(f"{BCRA_BASE_URL}?{urlencode({'limit': 2000, 'offset': 0})}")
    return {int(row["idVariable"]): row for row in payload.get("results", []) if row.get("idVariable") is not None}


def write_xlsx(path: Path, frame: pd.DataFrame, metadata: pd.DataFrame) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        frame.to_excel(writer, sheet_name="BCRA_public_series", index=False)
        metadata.to_excel(writer, sheet_name="Metadata", index=False)

    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        for column_cells in ws.columns:
            width = min(max(len(str(c.value)) if c.value is not None else 0 for c in column_cells) + 2, 55)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width
        ws.auto_filter.ref = ws.dimensions
    wb.save(path)


def update_snapshot_manifest(root: Path, output_dir: Path, csv_path: Path, xlsx_path: Path, rows: int) -> None:
    manifest_path = output_dir / "manifest.json"
    if not manifest_path.exists():
        return
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest.setdefault("files", {})
    manifest.setdefault("counts", {})
    manifest["files"]["bcra_public_series_csv"] = relpath(root, csv_path)
    manifest["files"]["bcra_public_series_xlsx"] = relpath(root, xlsx_path)
    manifest["counts"]["bcra_public_series"] = rows
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    latest_path = output_dir.parent / "latest.json"
    if latest_path.exists():
        latest = json.loads(latest_path.read_text(encoding="utf-8"))
        if latest.get("snapshot_dir") == relpath(root, output_dir):
            latest.update({k: v for k, v in manifest.items() if k in {"files", "counts", "warnings"}})
            latest_path.write_text(json.dumps(latest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def fetch_json(url: str) -> dict[str, Any]:
    request = Request(url, headers={"Accept": "application/json", "User-Agent": "rendimientos-ar-etl/1.0"})
    with urlopen(request, timeout=60) as response:
        return json.loads(response.read().decode("utf-8"))


def relpath(root: Path, path: Path) -> str:
    return str(path.resolve().relative_to(root.resolve())).replace("\\", "/")


if __name__ == "__main__":
    main()
