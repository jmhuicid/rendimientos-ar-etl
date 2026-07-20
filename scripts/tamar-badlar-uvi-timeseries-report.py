from __future__ import annotations

import argparse
import json
import math
from datetime import date, datetime
from pathlib import Path
from urllib.request import Request, urlopen

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


BCRA_VARIABLES = [
    {"id": 7, "key": "BADLAR", "label": "BADLAR privados", "unit": "% TNA"},
    {"id": 44, "key": "TAMAR", "label": "TAMAR privados", "unit": "% TNA"},
    {"id": 31, "key": "UVI_UVA", "label": "UVI/UVA", "unit": "indice"},
]

DATA912_ENDPOINTS = {
    "letras": "https://data912.com/live/arg_notes",
    "titulos_publicos": "https://data912.com/live/arg_bonds",
}

DEFAULT_LETRAS = [
    "S17A6",
    "S30A6",
    "S15Y6",
    "S29Y6",
    "T30J6",
    "S31L6",
    "S31G6",
    "S30S6",
    "S30O6",
    "S30N6",
    "T15E7",
    "T30A7",
    "T31Y7",
    "T30J7",
]

DEFAULT_TITULOS = [
    "AL29",
    "GD29",
    "AL30",
    "GD30",
    "AL35",
    "GD35",
    "AE38",
    "GD38",
    "AL41",
    "GD41",
    "AO27",
    "AN29",
    "BPD7",
    "TZX26",
    "TZXO6",
    "TX26",
    "TZXD6",
    "TZXM7",
    "TZX27",
    "TZXD7",
    "TZX28",
    "TX28",
    "DICP",
    "PARP",
]


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Genera un analisis comparativo TAMAR/BADLAR/UVI-UVA con letras y titulos publicos."
    )
    parser.add_argument("--input", default=None, help="Excel historico BYMA. Si se omite, busca data/input/series_historicas_YYYYMMDD.xlsx.")
    parser.add_argument("--output-dir", default=None, help="Directorio de salida. Default: data/reports.")
    parser.add_argument(
        "--config",
        default=None,
        help="Config de rendimientos-ar-etl para tickers de letras/titulos.",
    )
    parser.add_argument("--today", default=date.today().isoformat(), help="Fecha de corte YYYY-MM-DD.")
    parser.add_argument("--start-date", default="2023-01-01", help="Fecha inicial para ejemplo inversor YYYY-MM-DD.")
    parser.add_argument("--capital", type=float, default=100_000_000, help="Capital inicial para ejemplo inversor.")
    parser.add_argument("--future-inflation-dec", type=float, default=1.0, help="Inflacion mensual supuesta para diciembre 2026.")
    parser.add_argument("--future-tamar-dec-tna", type=float, default=22.1, help="TAMAR TNA diciembre 2026 segun REM o supuesto.")
    args = parser.parse_args()

    today = pd.to_datetime(args.today).date()
    start_date = pd.to_datetime(args.start_date).date()
    root = Path(__file__).resolve().parents[1]
    input_path = resolve_input_path(root, today, args.input)
    out_root = Path(args.output_dir) if args.output_dir else root / "data" / "reports"
    out_dir = out_root / today.isoformat()
    out_dir.mkdir(parents=True, exist_ok=True)

    config_path = Path(args.config) if args.config else root / "public" / "config.json"
    letras_cfg, titulos_cfg = load_reference_tickers(config_path)
    especies, indices = load_byma_workbook(input_path)
    bcra_raw = collect_bcra(today)
    bcra_panel, bcra_norm, bcra_rates, bcra_summary = build_bcra_tables(bcra_raw, today)
    investor_example = build_investor_example(bcra_panel, args.capital, start_date, today)
    future_scenarios = build_future_scenarios(
        bcra_panel,
        args.capital,
        today,
        args.future_inflation_dec,
        args.future_tamar_dec_tna,
    )
    trend_corr = build_trend_correlations(bcra_rates, start_date)
    live = collect_live_public_instruments(letras_cfg, titulos_cfg)
    byma_filtered = filter_byma_public_instruments(especies, letras_cfg, titulos_cfg)
    byma_summary = build_byma_summary(especies, letras_cfg, titulos_cfg)
    comparative = build_comparative(bcra_summary, bcra_rates, byma_summary, live, input_path, today)

    output_xlsx = out_dir / f"analisis_tamar_badlar_uvi_titulos_{today:%Y%m%d}.xlsx"
    output_md = out_dir / f"analisis_tamar_badlar_uvi_titulos_{today:%Y%m%d}.md"
    output_xlsx = available_output_path(output_xlsx)
    output_md = available_output_path(output_md)

    write_workbook(
        output_xlsx,
        {
            "Resumen": comparative,
            "BCRA_series": bcra_panel,
            "BCRA_base100": bcra_norm,
            "Tasas_comparables": bcra_rates,
            "Ejemplo_100M_2023": investor_example,
            "Escenarios_futuros_2026": future_scenarios,
            "Correlaciones_2023": trend_corr,
            "BCRA_ultimos": bcra_summary,
            "BYMA_titulos_resumen": byma_summary,
            "BYMA_titulos_historico": byma_filtered,
            "data912_vivo": live,
            "BYMA_indices": indices,
        },
    )
    write_markdown(output_md, comparative, bcra_summary, bcra_rates, investor_example, future_scenarios, trend_corr, byma_summary, live, today)

    print(f"OK Excel: {output_xlsx}")
    print(f"OK Markdown: {output_md}")


def load_reference_tickers(config_path: Path) -> tuple[list[str], list[str]]:
    letras = list(DEFAULT_LETRAS)
    titulos = list(DEFAULT_TITULOS)
    if config_path.exists():
        cfg = json.loads(config_path.read_text(encoding="utf-8"))
        letras = [
            item["ticker"]
            for item in cfg.get("lecaps", {}).get("letras", [])
            if item.get("activo") and item.get("ticker")
        ] or letras
        titulos = sorted(
            set(cfg.get("soberanos", {}).keys())
            | set(cfg.get("bonos_cer", {}).keys())
            | set(titulos)
        )
    return letras, titulos


def resolve_input_path(root: Path, today: date, explicit_input: str | None) -> Path:
    if explicit_input:
        return Path(explicit_input)
    expected = root / "data" / "input" / f"series_historicas_{today:%Y%m%d}.xlsx"
    if expected.exists():
        return expected
    candidates = sorted((root / "data" / "input").glob("series_historicas_*.xlsx"))
    if candidates:
        return candidates[-1]
    return expected


def available_output_path(path: Path) -> Path:
    if not path.exists():
        return path
    try:
        with path.open("a"):
            return path
    except PermissionError:
        stamp = datetime.now().strftime("%H%M%S")
        return path.with_name(f"{path.stem}_{stamp}{path.suffix}")


def load_byma_workbook(path: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    especies = pd.read_excel(path, sheet_name="Especies")
    indices = pd.read_excel(path, sheet_name="Indices")
    especies["FECHA"] = pd.to_datetime(especies["FECHA"], errors="coerce")
    indices["FECHA"] = pd.to_datetime(indices["FECHA"], dayfirst=True, errors="coerce")
    for col in [
        "APERTURA",
        "MAXIMO",
        "MINIMO",
        "CIERRE",
        "PRECIO PROMEDIO",
        "VOLUMEN NOMINAL",
        "MONTO NEGOCIADO",
        "CANTIDAD DE OPERACIONES",
    ]:
        if col in especies.columns:
            especies[col] = pd.to_numeric(especies[col], errors="coerce")
    for col in ["APERTURA", "MINIMO", "MAXIMO", "ULTIMO"]:
        if col in indices.columns:
            indices[col] = pd.to_numeric(indices[col], errors="coerce")
    return especies, indices


def collect_bcra(today: date) -> pd.DataFrame:
    frames = []
    for item in BCRA_VARIABLES:
        detalle = fetch_bcra_detail(item["id"])
        frame = pd.DataFrame(detalle)
        if frame.empty:
            continue
        frame["fecha"] = pd.to_datetime(frame["fecha"], errors="coerce")
        frame["valor"] = pd.to_numeric(frame["valor"], errors="coerce")
        frame = frame.dropna(subset=["fecha", "valor"])
        frame = frame[frame["fecha"].dt.date <= today]
        frame["serie"] = item["key"]
        frame["nombre"] = item["label"]
        frame["unidad"] = item["unit"]
        frames.append(frame[["fecha", "serie", "nombre", "unidad", "valor"]])
    return pd.concat(frames, ignore_index=True).sort_values(["serie", "fecha"])


def fetch_bcra_detail(variable_id: int) -> list[dict]:
    limit = 1000
    offset = 0
    rows: list[dict] = []
    total = None
    while total is None or offset < total:
        url = f"https://api.bcra.gob.ar/estadisticas/v4.0/Monetarias/{variable_id}?limit={limit}&offset={offset}"
        payload = fetch_json(url)
        total = payload.get("metadata", {}).get("resultset", {}).get("count", 0)
        detail = payload.get("results", [{}])[0].get("detalle", [])
        if not detail:
            break
        rows.extend(detail)
        offset += limit
    return rows


def build_bcra_tables(raw: pd.DataFrame, today: date) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    panel = raw.pivot_table(index="fecha", columns="serie", values="valor", aggfunc="last").sort_index()
    panel = panel.ffill()
    panel = panel[panel.index.date <= today]

    last_year_start = pd.Timestamp(today) - pd.Timedelta(days=365)
    window = panel[panel.index >= last_year_start].dropna(how="all")
    norm = window.copy()
    for col in norm.columns:
        valid = norm[col].dropna()
        if valid.empty:
            continue
        norm[col] = norm[col] / valid.iloc[0] * 100
    norm = norm.reset_index().rename(columns={"fecha": "FECHA"})
    panel_out = panel.reset_index().rename(columns={"fecha": "FECHA"})
    rates = build_comparable_rates(panel)

    rows = []
    for item in BCRA_VARIABLES:
        serie = item["key"]
        s = panel[serie].dropna() if serie in panel else pd.Series(dtype=float)
        if s.empty:
            continue
        latest = s.iloc[-1]
        previous = s.iloc[-2] if len(s) > 1 else math.nan
        m1 = value_on_or_before(s, s.index[-1] - pd.Timedelta(days=30))
        y1 = value_on_or_before(s, s.index[-1] - pd.Timedelta(days=365))
        rows.append(
            {
                "serie": serie,
                "nombre": item["label"],
                "unidad": item["unit"],
                "ultima_fecha": s.index[-1].date().isoformat(),
                "ultimo_valor": latest,
                "var_abs_dato_previo": latest - previous if pd.notna(previous) else None,
                "var_pct_30d": pct_change(latest, m1),
                "var_pct_1y": pct_change(latest, y1),
            }
        )
    return panel_out, norm, rates, pd.DataFrame(rows)


def build_comparable_rates(panel: pd.DataFrame) -> pd.DataFrame:
    rates = pd.DataFrame(index=panel.index)
    if "BADLAR" in panel:
        rates["BADLAR_TEA"] = tna_to_tea(panel["BADLAR"])
    if "TAMAR" in panel:
        rates["TAMAR_TEA"] = tna_to_tea(panel["TAMAR"])
    if "UVI_UVA" in panel:
        uvi = panel["UVI_UVA"].dropna()
        rates["UVI_TEA_30D"] = rolling_index_tea(uvi, 30).reindex(panel.index)
        rates["UVI_TEA_90D"] = rolling_index_tea(uvi, 90).reindex(panel.index)
        rates["UVI_TEA_365D"] = rolling_index_tea(uvi, 365).reindex(panel.index)
    return rates.dropna(how="all").reset_index().rename(columns={"fecha": "FECHA"})


def build_investor_example(panel_out: pd.DataFrame, capital: float, start_date: date, today: date) -> pd.DataFrame:
    panel = panel_out.copy()
    panel["FECHA"] = pd.to_datetime(panel["FECHA"])
    panel = panel.set_index("FECHA").sort_index()
    end_ts = pd.Timestamp(today)

    long_start = pd.Timestamp(start_date)
    tamar_start = first_valid_date(panel, "TAMAR", long_start, end_ts)
    rows = []
    rows.extend(
        investor_scenario_rows(
            panel,
            "Desde 2023: BADLAR vs UVI/UVA",
            capital,
            long_start,
            end_ts,
            [("BADLAR", "BADLAR privados"), ("UVI_UVA", "UVI/UVA")],
        )
    )
    if tamar_start is not None:
        rows.extend(
            investor_scenario_rows(
                panel,
                "Desde inicio TAMAR: mismas fechas",
                capital,
                tamar_start,
                end_ts,
                [("BADLAR", "BADLAR privados"), ("TAMAR", "TAMAR privados"), ("UVI_UVA", "UVI/UVA")],
            )
        )
    rows.extend(
        investor_scenario_rows(
            panel,
            "Desde 2026: mismas fechas",
            capital,
            pd.Timestamp("2026-01-01"),
            end_ts,
            [("BADLAR", "BADLAR privados"), ("TAMAR", "TAMAR privados"), ("UVI_UVA", "UVI/UVA")],
        )
    )
    return pd.DataFrame(rows)


def first_valid_date(panel: pd.DataFrame, column: str, start_ts: pd.Timestamp, end_ts: pd.Timestamp) -> pd.Timestamp | None:
    if column not in panel:
        return None
    valid = panel.loc[(panel.index >= start_ts) & (panel.index <= end_ts), column].dropna()
    if valid.empty:
        return None
    return valid.index[0]


def investor_scenario_rows(
    panel: pd.DataFrame,
    scenario: str,
    capital: float,
    start_ts: pd.Timestamp,
    end_ts: pd.Timestamp,
    series: list[tuple[str, str]],
) -> list[dict]:
    window = panel[(panel.index >= start_ts) & (panel.index <= end_ts)].ffill()
    if window.empty:
        return []

    rows = []
    for serie, label in series:
        if serie not in window:
            continue
        values = window[serie].dropna()
        if values.empty:
            continue
        if serie == "UVI_UVA":
            if values.iloc[0] <= 0:
                continue
            final_capital = capital * (values.iloc[-1] / values.iloc[0])
            method = "Ajuste por indice"
        else:
            factors = 1 + (values / 100) / 365
            final_capital = capital * factors.prod()
            method = "TNA reinvertida diaria"
        rows.append(investor_row(scenario, label, method, capital, final_capital, values.index[0], values.index[-1]))

    return rows


def investor_row(scenario: str, name: str, method: str, initial: float, final: float, start: pd.Timestamp, end: pd.Timestamp) -> dict:
    days = max((end - start).days, 1)
    gain = final - initial
    return {
        "escenario": scenario,
        "variable": name,
        "metodo": method,
        "fecha_inicio": start.date().isoformat(),
        "fecha_final": end.date().isoformat(),
        "capital_inicial": initial,
        "capital_final": final,
        "ganancia_nominal": gain,
        "multiplicador": final / initial,
        "rendimiento_acumulado_pct": (final / initial - 1) * 100,
        "tea_equivalente_periodo_pct": ((final / initial) ** (365 / days) - 1) * 100,
    }


def build_future_scenarios(
    panel_out: pd.DataFrame,
    capital: float,
    today: date,
    inflation_dec_pct: float,
    tamar_dec_tna: float,
) -> pd.DataFrame:
    panel = panel_out.copy()
    panel["FECHA"] = pd.to_datetime(panel["FECHA"])
    panel = panel.set_index("FECHA").sort_index().ffill()
    today_ts = pd.Timestamp(today)
    end_ts = pd.Timestamp("2026-12-31")
    if today_ts >= end_ts or panel.empty:
        return pd.DataFrame()

    latest = panel.loc[panel.index <= today_ts].iloc[-1]
    months = pd.date_range(today_ts + pd.offsets.MonthEnd(1), end_ts, freq="ME")
    if months.empty:
        return pd.DataFrame()

    current_uvi_30d = observed_monthly_index_change(panel["UVI_UVA"].dropna(), today_ts)
    monthly_inflation = linear_path(current_uvi_30d, inflation_dec_pct, len(months))
    tamar_path = linear_path(float(latest.get("TAMAR", tamar_dec_tna)), tamar_dec_tna, len(months))
    spread_badlar_tamar = float(latest.get("BADLAR", tamar_dec_tna) - latest.get("TAMAR", tamar_dec_tna))
    badlar_path = [max(t + spread_badlar_tamar, 0) for t in tamar_path]

    rows = []
    rows.append(projected_rate_row("Jul-Dic 2026 estimado", "BADLAR privados", "TNA proyectada, reinversion diaria", capital, badlar_path))
    rows.append(projected_rate_row("Jul-Dic 2026 estimado", "TAMAR privados", "REM mayo 2026: converge a 22,1% TNA en diciembre", capital, tamar_path))
    rows.append(projected_index_row("Jul-Dic 2026 estimado", "UVI/UVA", "Inflacion mensual converge a 1,0% en diciembre", capital, monthly_inflation))
    for row in rows:
        row["fecha_inicio"] = (today_ts + pd.Timedelta(days=1)).date().isoformat()
        row["fecha_final"] = end_ts.date().isoformat()
        row["inflacion_mensual_dic_2026_supuesto"] = inflation_dec_pct
        row["tamar_dic_2026_tna_rem"] = tamar_dec_tna
    return pd.DataFrame(rows)


def observed_monthly_index_change(series: pd.Series, today_ts: pd.Timestamp) -> float:
    current = value_on_or_before(series, today_ts)
    previous = value_on_or_before(series, today_ts - pd.Timedelta(days=30))
    change = pct_change(current, previous)
    return float(change) if change is not None and pd.notna(change) else 1.0


def linear_path(start_value: float, end_value: float, steps: int) -> list[float]:
    if steps <= 1:
        return [end_value]
    return [start_value + (end_value - start_value) * i / (steps - 1) for i in range(steps)]


def projected_rate_row(scenario: str, name: str, method: str, capital: float, monthly_tna: list[float]) -> dict:
    final = capital
    for tna in monthly_tna:
        final *= (1 + (tna / 100) / 365) ** 30
    return projected_row(scenario, name, method, capital, final)


def projected_index_row(scenario: str, name: str, method: str, capital: float, monthly_inflation: list[float]) -> dict:
    final = capital
    for monthly in monthly_inflation:
        final *= 1 + monthly / 100
    return projected_row(scenario, name, method, capital, final)


def projected_row(scenario: str, name: str, method: str, initial: float, final: float) -> dict:
    return {
        "escenario": scenario,
        "variable": name,
        "metodo": method,
        "capital_inicial": initial,
        "capital_final_estimado": final,
        "ganancia_nominal_estimada": final - initial,
        "multiplicador_estimado": final / initial,
        "rendimiento_estimado_pct": (final / initial - 1) * 100,
    }


def build_trend_correlations(bcra_rates: pd.DataFrame, start_date: date) -> pd.DataFrame:
    if bcra_rates.empty:
        return pd.DataFrame()
    rates = bcra_rates.copy()
    rates["FECHA"] = pd.to_datetime(rates["FECHA"])
    rates = rates[rates["FECHA"].dt.date >= start_date].set_index("FECHA").sort_index()
    cols = [c for c in ["BADLAR_TEA", "TAMAR_TEA", "UVI_TEA_30D", "UVI_TEA_90D", "UVI_TEA_365D"] if c in rates]
    if len(cols) < 2:
        return pd.DataFrame()

    level_corr = rates[cols].corr()
    change_corr = rates[cols].diff(30).corr()
    rows = []
    for a in cols:
        for b in cols:
            if a >= b:
                continue
            rows.append(
                {
                    "serie_a": a,
                    "serie_b": b,
                    "correlacion_nivel": level_corr.loc[a, b],
                    "correlacion_cambio_30d": change_corr.loc[a, b],
                    "lectura": correlation_reading(level_corr.loc[a, b], change_corr.loc[a, b]),
                }
            )
    return pd.DataFrame(rows)


def correlation_reading(level_corr: float, change_corr: float) -> str:
    if pd.isna(level_corr):
        return "Datos insuficientes."
    if level_corr >= 0.7:
        base = "Tendencias muy alineadas"
    elif level_corr >= 0.35:
        base = "Tendencias moderadamente alineadas"
    elif level_corr <= -0.35:
        base = "Tendencias inversas"
    else:
        base = "Relacion debil en niveles"
    if pd.notna(change_corr) and abs(change_corr) < 0.25:
        return f"{base}; los cambios de corto plazo no necesariamente se mueven juntos."
    return f"{base}."


def tna_to_tea(series: pd.Series) -> pd.Series:
    return ((1 + (series / 100) / 365) ** 365 - 1) * 100


def rolling_index_tea(series: pd.Series, window_days: int) -> pd.Series:
    values = []
    for current_date, current_value in series.items():
        previous_date = current_date - pd.Timedelta(days=window_days)
        previous_value = value_on_or_before(series, previous_date)
        if previous_value is None or previous_value <= 0:
            values.append(None)
            continue
        actual_previous_date = series[series.index <= previous_date].dropna().index[-1]
        days = max((current_date - actual_previous_date).days, 1)
        tea = ((current_value / previous_value) ** (365 / days) - 1) * 100
        values.append(tea)
    return pd.Series(values, index=series.index)


def collect_live_public_instruments(letras_cfg: list[str], titulos_cfg: list[str]) -> pd.DataFrame:
    rows = []
    for group, url in DATA912_ENDPOINTS.items():
        data = fetch_json(url)
        configured = set(letras_cfg if group == "letras" else titulos_cfg)
        for item in data if isinstance(data, list) else []:
            symbol = str(item.get("symbol", "")).strip()
            base = symbol[:-1] if symbol.endswith(("D", "C", "X", "Y", "Z")) else symbol
            is_selected = symbol in configured or base in configured
            rows.append(
                {
                    "grupo": group,
                    "symbol": symbol,
                    "base_symbol": base,
                    "seleccion_rendimientos_ar": is_selected,
                    "ultimo": number_or_none(item.get("c")),
                    "bid": number_or_none(item.get("px_bid")),
                    "ask": number_or_none(item.get("px_ask")),
                    "volumen": number_or_none(item.get("v")),
                    "operaciones": number_or_none(item.get("q_op")),
                    "pct_change": number_or_none(item.get("pct_change")),
                    "fuente": url,
                }
            )
    frame = pd.DataFrame(rows)
    if frame.empty:
        return frame
    return frame.sort_values(["seleccion_rendimientos_ar", "grupo", "symbol"], ascending=[False, True, True])


def build_byma_summary(especies: pd.DataFrame, letras_cfg: list[str], titulos_cfg: list[str]) -> pd.DataFrame:
    df = filter_byma_public_instruments(especies, letras_cfg, titulos_cfg)
    if df.empty:
        return pd.DataFrame()

    rows = []
    for symbol, group in df.sort_values("FECHA").groupby("SIMBOLO"):
        valid = group.dropna(subset=["FECHA", "CIERRE"])
        if valid.empty:
            continue
        latest = valid.iloc[-1]
        first = valid.iloc[0]
        prev = valid.iloc[-2] if len(valid) > 1 else None
        rows.append(
            {
                "SIMBOLO": symbol,
                "base_symbol": latest["base_symbol"],
                "primera_fecha_archivo": first["FECHA"].date().isoformat(),
                "ultima_fecha_archivo": latest["FECHA"].date().isoformat(),
                "ruedas": len(valid),
                "cierre_inicial": first["CIERRE"],
                "cierre_ultimo": latest["CIERRE"],
                "var_pct_archivo": pct_change(latest["CIERRE"], first["CIERRE"]),
                "var_pct_1rueda": pct_change(latest["CIERRE"], prev["CIERRE"]) if prev is not None else None,
                "volumen_nominal_ultimo": latest.get("VOLUMEN NOMINAL"),
                "monto_negociado_ultimo": latest.get("MONTO NEGOCIADO"),
                "operaciones_ultimo": latest.get("CANTIDAD DE OPERACIONES"),
            }
        )
    return pd.DataFrame(rows).sort_values(["ultima_fecha_archivo", "SIMBOLO"], ascending=[False, True])


def filter_byma_public_instruments(
    especies: pd.DataFrame, letras_cfg: list[str], titulos_cfg: list[str]
) -> pd.DataFrame:
    selected = set(letras_cfg) | set(titulos_cfg)
    df = especies.copy()
    df["base_symbol"] = df["SIMBOLO"].astype(str).str.replace(r"[DCXYZ]$", "", regex=True)
    df = df[(df["SIMBOLO"].isin(selected)) | (df["base_symbol"].isin(selected))]
    return df.sort_values(["SIMBOLO", "FECHA"])


def build_comparative(
    bcra_summary: pd.DataFrame,
    bcra_rates: pd.DataFrame,
    byma_summary: pd.DataFrame,
    live: pd.DataFrame,
    input_path: Path,
    today: date,
) -> pd.DataFrame:
    byma_last = None
    byma_first = None
    if not byma_summary.empty:
        byma_last = byma_summary["ultima_fecha_archivo"].max()
        byma_first = byma_summary["primera_fecha_archivo"].min()

    selected_live = live[live.get("seleccion_rendimientos_ar", False) == True] if not live.empty else pd.DataFrame()
    latest_rates = bcra_rates.dropna(how="all").iloc[-1].to_dict() if not bcra_rates.empty else {}
    rows = [
        {
            "bloque": "Corte",
            "metrica": "Fecha de generacion",
            "valor": today.isoformat(),
            "lectura": "Reporte generado con datos disponibles al momento de corrida.",
        },
        {
            "bloque": "BYMA historico",
            "metrica": "Archivo fuente",
            "valor": input_path.name,
            "lectura": f"Rango detectado {byma_first or 'n/d'} a {byma_last or 'n/d'}. Si el archivo descargado hoy trae maximo 2025-06-27, el historico BYMA no esta actualizado a 2026-06-30.",
        },
        {
            "bloque": "BCRA",
            "metrica": "Series oficiales",
            "valor": ", ".join(bcra_summary["serie"].astype(str)) if not bcra_summary.empty else "n/d",
            "lectura": "BADLAR y TAMAR se convierten de TNA a TEA. UVI/UVA se convierte de indice a TEA anualizada por ventanas moviles.",
        },
        {
            "bloque": "Tasas comparables",
            "metrica": "UVI/UVA TEA 30D",
            "valor": latest_rates.get("UVI_TEA_30D", "n/d"),
            "lectura": "Tasa efectiva anualizada de la variacion observada del indice en los ultimos 30 dias calendario disponibles.",
        },
        {
            "bloque": "Tasas comparables",
            "metrica": "BADLAR TEA / TAMAR TEA",
            "valor": f"{fmt_pct(latest_rates.get('BADLAR_TEA'))} / {fmt_pct(latest_rates.get('TAMAR_TEA'))}",
            "lectura": "Conversion de TNA a TEA con capitalizacion diaria para comparar en la misma escala que UVI/UVA anualizado.",
        },
        {
            "bloque": "Mercado vivo",
            "metrica": "Instrumentos data912 seleccionados",
            "valor": len(selected_live),
            "lectura": "Snapshot vivo operativo para Letras y Titulos Publicos tomado de la misma fuente usada en rendimientos-ar-etl.",
        },
    ]
    for _, row in bcra_summary.iterrows():
        rows.append(
            {
                "bloque": "BCRA ultimo",
                "metrica": row["nombre"],
                "valor": row["ultimo_valor"],
                "lectura": f"{row['unidad']} al {row['ultima_fecha']}; var. 30d {fmt_pct(row['var_pct_30d'])}, var. 1y {fmt_pct(row['var_pct_1y'])}.",
            }
        )
    return pd.DataFrame(rows)


def write_workbook(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, frame in sheets.items():
            safe = name[:31]
            frame.to_excel(writer, sheet_name=safe, index=False)

    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        for column_cells in ws.columns:
            width = min(max(len(str(c.value)) if c.value is not None else 0 for c in column_cells) + 2, 45)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width
        ws.auto_filter.ref = ws.dimensions

    if "BCRA_base100" in wb.sheetnames:
        ws = wb["BCRA_base100"]
        if ws.max_row > 2 and ws.max_column > 2:
            chart = LineChart()
            chart.title = "BADLAR vs TAMAR vs UVI/UVA - base 100"
            chart.y_axis.title = "Base 100"
            chart.x_axis.title = "Fecha"
            data = Reference(ws, min_col=2, max_col=ws.max_column, min_row=1, max_row=ws.max_row)
            cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 9
            chart.width = 18
            ws.add_chart(chart, "F2")

    if "Tasas_comparables" in wb.sheetnames:
        ws = wb["Tasas_comparables"]
        if ws.max_row > 2 and ws.max_column > 2:
            chart = LineChart()
            chart.title = "Tasas comparables - TEA"
            chart.y_axis.title = "% TEA"
            chart.x_axis.title = "Fecha"
            data = Reference(ws, min_col=2, max_col=ws.max_column, min_row=1, max_row=ws.max_row)
            cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 9
            chart.width = 18
            ws.add_chart(chart, "H2")

    wb.save(path)


def write_markdown(
    path: Path,
    comparative: pd.DataFrame,
    bcra_summary: pd.DataFrame,
    bcra_rates: pd.DataFrame,
    investor_example: pd.DataFrame,
    future_scenarios: pd.DataFrame,
    trend_corr: pd.DataFrame,
    byma_summary: pd.DataFrame,
    live: pd.DataFrame,
    today: date,
) -> None:
    lines = [
        "# Analisis comparativo TAMAR vs BADLAR vs UVI/UVA",
        "",
        f"Fecha de generacion: {today.isoformat()}",
        "",
        "## Lectura rapida",
        "",
    ]
    for _, row in comparative.iterrows():
        lines.append(f"- **{row['bloque']} / {row['metrica']}**: {row['valor']}. {row['lectura']}")
    lines.extend(
        [
            "",
            "## Que mide cada variable",
            "",
            "- **BADLAR**: tasa promedio pagada por bancos privados por depositos a plazo fijo mayoristas. Desde el lado inversor, aproxima una referencia de tasa pasiva bancaria para pesos.",
            "- **TAMAR**: tasa promedio para depositos a plazo fijo mayoristas de mayor monto. Suele mirar un segmento institucional y puede quedar por encima o por debajo de BADLAR segun liquidez bancaria.",
            "- **UVI/UVA**: indice que ajusta por inflacion. Por eso su TEA anualizada suele verse mas estable que una tasa de mercado: no es una tasa ofrecida, sino la capitalizacion efectiva de la inflacion observada en el periodo.",
            "",
            "Desde el inversor, la pregunta no es solo que serie sube o baja, sino si la tasa nominal compensa el ritmo del indice inflacionario. Cuando UVI/UVA anualizado supera BADLAR/TAMAR TEA, el deposito remunerado queda corriendo de atras a la inflacion del periodo observado.",
        ]
    )
    lines.extend(["", "## Ultimos datos BCRA", ""])
    if not bcra_summary.empty:
        lines.extend(markdown_table(bcra_summary))
    lines.extend(["", "## Tasas comparables TEA", ""])
    if not bcra_rates.empty:
        latest = bcra_rates.tail(1).copy()
        lines.extend(markdown_table(latest))
        lines.extend(
            [
                "",
                "Metodologia: BADLAR/TAMAR se transforman desde TNA a TEA con capitalizacion diaria. UVI/UVA se transforma desde indice a tasa efectiva anualizada: `((indice_actual / indice_previo) ** (365 / dias) - 1) * 100`.",
            ]
        )
    lines.extend(["", "## Ejemplo inversor", ""])
    if not investor_example.empty:
        lines.append(
            "Supuesto: capital inicial de $100.000.000, reinvirtiendo tasas diariamente con TNA/365 y ajustando UVI/UVA por variacion del indice."
        )
        lines.append(
            "Escenario 1: desde 2023 se compara BADLAR contra UVI/UVA, porque TAMAR no tiene dato disponible desde esa fecha."
        )
        lines.append(
            "Escenario 2: desde el inicio disponible de TAMAR se reinician los $100.000.000 y se comparan BADLAR, TAMAR y UVI/UVA en las mismas fechas."
        )
        lines.append(
            "Escenario 3: desde 2026-01-01 se vuelve a reiniciar el capital para mirar el ano corriente con las tres variables en paralelo."
        )
        lines.append("")
        display = investor_example.copy()
        for col in ["capital_inicial", "capital_final", "ganancia_nominal"]:
            display[col] = display[col].map(format_ars)
        for col in ["rendimiento_acumulado_pct", "tea_equivalente_periodo_pct"]:
            display[col] = display[col].map(fmt_pct)
        display["multiplicador"] = display["multiplicador"].map(lambda v: f"{v:.2f}x")
        lines.extend(markdown_table(display))
    lines.extend(["", "## Escenario futuro a diciembre 2026", ""])
    lines.extend(
        [
            "Para proyectar hacia adelante ya no usamos dato observado puro: se arma un estimativo. La idea es suponer un proceso de desinflacion donde la inflacion mensual converge gradualmente hacia 1,0% en diciembre 2026, y tomar como ancla de mercado la TAMAR esperada por REM para diciembre 2026 en torno a 22,1% TNA.",
            "",
            "Si tasas e inflacion bajan juntas, la lectura del inversor cambia: ya no alcanza con mirar quien gana en el acumulado historico, sino si la tasa futura queda por encima o por debajo del sendero de inflacion esperado. En un escenario de inflacion bajando a 1% mensual, una TAMAR/BADLAR que no baje demasiado rapido puede volver a verse competitiva en terminos reales de corto plazo.",
        ]
    )
    if not future_scenarios.empty:
        future_display = future_scenarios.copy()
        for col in ["capital_inicial", "capital_final_estimado", "ganancia_nominal_estimada"]:
            future_display[col] = future_display[col].map(format_ars)
        for col in ["rendimiento_estimado_pct", "inflacion_mensual_dic_2026_supuesto", "tamar_dic_2026_tna_rem"]:
            future_display[col] = future_display[col].map(fmt_pct)
        future_display["multiplicador_estimado"] = future_display["multiplicador_estimado"].map(lambda v: f"{v:.2f}x")
        lines.append("")
        lines.extend(markdown_table(future_display))
        lines.extend(
            [
                "",
                "Fuente/criterio: REM BCRA publicado en junio 2026 informa TAMAR proyectada de 22,1% TNA para diciembre 2026; el sendero de inflacion a 1,0% mensual en diciembre es un supuesto de desinflacion para sensibilizar el ejemplo.",
            ]
        )
    lines.extend(["", "## Correlacion de tendencias desde 2023", ""])
    if not trend_corr.empty:
        corr_display = trend_corr.copy()
        for col in ["correlacion_nivel", "correlacion_cambio_30d"]:
            corr_display[col] = corr_display[col].map(lambda v: "n/d" if pd.isna(v) else f"{v:.2f}")
        lines.extend(markdown_table(corr_display))
        lines.extend(
            [
                "",
                "Lectura: la correlacion en niveles muestra si las curvas tienden a moverse en la misma direccion durante el periodo. La correlacion de cambios a 30 dias es mas exigente: pregunta si los giros de corto plazo tambien coinciden.",
            ]
        )
    lines.extend(["", "## Instrumentos publicos", ""])
    lines.append(
        f"- BYMA historico filtrado: {len(byma_summary)} series de Letras/Titulos detectadas."
    )
    lines.append(
        f"- data912 vivo: {len(live)} instrumentos totales; "
        f"{int(live['seleccion_rendimientos_ar'].sum()) if not live.empty else 0} coinciden con la seleccion de rendimientos-ar-etl."
    )
    lines.extend(
        [
            "",
            "Nota: el pedido menciona UVI; la API usada por rendimientos-ar-etl expone el ID 31 como UVA. Se etiqueta UVI/UVA para mantener trazabilidad del pedido y de la fuente.",
            "",
        ]
    )
    path.write_text("\n".join(lines), encoding="utf-8")


def markdown_table(frame: pd.DataFrame) -> list[str]:
    cols = list(frame.columns)
    lines = [
        "| " + " | ".join(cols) + " |",
        "| " + " | ".join("---" for _ in cols) + " |",
    ]
    for _, row in frame.iterrows():
        values = [str(row[col]) if pd.notna(row[col]) else "" for col in cols]
        values = [value.replace("|", "\\|") for value in values]
        lines.append("| " + " | ".join(values) + " |")
    return lines


def fetch_json(url: str):
    req = Request(url, headers={"Accept": "application/json", "User-Agent": "Serie_Tiempo/1.0"})
    with urlopen(req, timeout=60) as response:
        return json.loads(response.read().decode("utf-8"))


def value_on_or_before(series: pd.Series, target: pd.Timestamp):
    prior = series[series.index <= target].dropna()
    return prior.iloc[-1] if not prior.empty else None


def pct_change(current, previous):
    if previous is None or pd.isna(previous) or previous == 0 or pd.isna(current):
        return None
    return (float(current) / float(previous) - 1) * 100


def number_or_none(value):
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return None
    return parsed if math.isfinite(parsed) else None


def fmt_pct(value) -> str:
    if value is None or pd.isna(value):
        return "n/d"
    return f"{float(value):.2f}%"


def format_ars(value) -> str:
    if value is None or pd.isna(value):
        return "n/d"
    return f"${float(value):,.0f}".replace(",", ".")


if __name__ == "__main__":
    main()
